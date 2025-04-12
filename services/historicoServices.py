from collections import defaultdict
from datetime import date, datetime

from models.ciclo import Ciclo
from models.sensoresAA import SensoresAA
from models.sensores import Sensores
from models.estadoCiclo import EstadoCiclo
from models.receta import Receta
from models.equipo import Equipo

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from io import BytesIO

import logging

logger = logging.getLogger("uvicorn")

def generar_reporte_productividad(id_equipo, fecha_inicio, fecha_fin, db):
    fecha_inicio = datetime.combine(fecha_inicio, datetime.min.time())
    fecha_fin = datetime.combine(fecha_fin, datetime.max.time())
    lista_ciclos: list[Ciclo] = [] 
    linea1 = [7,8,9,10]
    linea2 = [11,12,13,14]

    nombre_maquina = ""
    def buscarReceta(id_receta):
        receta = (
            db.query(Receta)
            .filter(id_receta == Receta.id)
            .first()
        )
        return receta.nombre
    
    def bucarNombreEquipo(id_maquina):
        maquina = (db.query(Equipo).filter(id_maquina == Equipo.id).first())
        return maquina.nombre

    if id_equipo == 15:
        lista_ciclos= (
            db.query(Ciclo)
            .filter(Ciclo.idEquipo.in_(linea1))
            .filter(Ciclo.fecha_fin.between(fecha_inicio, fecha_fin))
            .all()
        )
        nombre_maquina = "Linea 1"
    if id_equipo == 16:
        lista_ciclos= (
            db.query(Ciclo)
            .filter(Ciclo.idEquipo.in_(linea2))
            .filter(Ciclo.fecha_fin.between(fecha_inicio, fecha_fin))
            .all()
        )
        nombre_maquina = "Linea 2"
    if id_equipo <= 14 and id_equipo != 0:
        lista_ciclos = (
            db.query(Ciclo)
            .filter(Ciclo.fecha_fin.between(fecha_inicio, fecha_fin))
            .filter(Ciclo.idEquipo == id_equipo)
            .all()
        )
        nombre_maquina = bucarNombreEquipo(id_equipo)
    if id_equipo == 0:
        lista_ciclos = (
            db.query(Ciclo)
            .filter(Ciclo.fecha_fin.between(fecha_inicio, fecha_fin))
            .all()
        )
        nombre_maquina = "Completo"

    #CONSTRUCCION DEL ARCHIVO XLMS
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Informe de productividad"
    logoPath = "cremona.png"
    img = Image(logoPath)
    img.width = 280
    img.height = 70
    sheet.add_image(img, 'D1')

    sheet.append(["Lista ciclos finalizados correctamente"])
    producto_cell = sheet.cell(row=sheet.max_row, column=1)
    producto_cell.font = Font(bold= True, size=20)

    sheet.append(["Fecha Inicio:", fecha_inicio.strftime("%Y-%m-%d")])
    fechaInicio_cell = sheet.cell(row=sheet.max_row, column=1)
    fechaInicio_cell.font = Font(bold=True, size=12)
    sheet.append(["Fecha Fin:", fecha_fin.strftime("%Y-%m-%d")])
    fechaFin_cell = sheet.cell(row=sheet.max_row, column=1)
    fechaFin_cell.font = Font(bold=True, size=12)

    sheet.append([f"Buscar: {nombre_maquina}"])
    producto_cell = sheet.cell(row=sheet.max_row, column=5)
    producto_cell.font = Font(bold= True, size=16)

    headers = ["ID_CICLO", "ESTADO_CICLO", "CANTIDAD_TORRE", "LOTE", "TIEMPO_TRANSCURRIDO", "FECHA_INICIO", "FECHA_FIN", "EQUIPO", "PESO","RECETA"]
    sheet.append(headers)
    header_fill = PatternFill(start_color="145f82", end_color="145f82", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)  

    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=sheet.max_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    start_row = sheet.max_row + 1

    resultado_fila = []

    for elem in lista_ciclos:
        item_equipo = bucarNombreEquipo(elem.idEquipo)
        item_receta = buscarReceta(elem.idReceta)
        resultado_fila.append([
            elem.id,
            elem.estadoMaquina,
            elem.cantidadTorres,
            elem.lote,
            elem.tiempoTranscurrido,
            elem.fecha_inicio,
            elem.fecha_fin,
            item_equipo,
            elem.peso,
            item_receta
        ])

    for item in resultado_fila:
        sheet.append(item)
    
    end_row = sheet.max_row
    start_col = 1
    end_col = len(headers)
    table_range = f"{sheet.cell(row=start_row -1, column=start_col).coordinate}:{sheet.cell(row=end_row, column=end_col).coordinate}"
    table_nombre = "GraficosHistorico"
    tabla = Table(displayName=table_nombre, ref=table_range)
    style = TableStyleInfo(showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tabla.tableStyleInfo = style

    # Agregar la tabla a la hoja
    sheet.add_table(tabla)
    sheet.append([])

    for col in sheet.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        sheet.column_dimensions[column_letter].width = max_length + 2

    excel_stream = BytesIO()
    workbook.save(excel_stream)
    workbook.close() 
    excel_stream.seek(0)  

    return excel_stream

def productividad_equipo(db, fecha_inicio, fecha_fin, id_equipo):
    fecha_inicio = datetime.combine(fecha_inicio, datetime.min.time())
    fecha_fin = datetime.combine(fecha_fin, datetime.max.time())

    respuesta = {}
    lista_receta = defaultdict(list)
    lista_ciclos: list[Ciclo] = [] 
    linea1 = [7,8,9,10]
    linea2 = [11,12,13,14]

    if id_equipo == 15:
        lista_ciclos= (
            db.query(Ciclo)
            .filter(Ciclo.idEquipo.in_(linea1))
            .filter(Ciclo.fecha_fin.between(fecha_inicio, fecha_fin))
            .all()
        )
    if id_equipo == 16:
        lista_ciclos= (
            db.query(Ciclo)
            .filter(Ciclo.idEquipo.in_(linea2))
            .filter(Ciclo.fecha_fin.between(fecha_inicio, fecha_fin))
            .all()
        )

    if id_equipo <= 14 and id_equipo != 0:
        lista_ciclos = (
            db.query(Ciclo)
            .filter(Ciclo.fecha_fin.between(fecha_inicio, fecha_fin))
            .filter(Ciclo.idEquipo == id_equipo)
            .all()
        )
    if id_equipo == 0:
        lista_ciclos = (
            db.query(Ciclo)
            .filter(Ciclo.fecha_fin.between(fecha_inicio, fecha_fin))
            .all()
        )

    total_peso = 0
    cantidad_ciclos_correctos = 0
    cantidad_ciclos_incorectos = 0

    for elem in lista_ciclos:
        total_peso+= elem.peso
        if elem.estadoMaquina == "Finalizado":
            cantidad_ciclos_correctos+= 1
        if elem.estadoMaquina == "Cancelado":
            cantidad_ciclos_incorectos+= 1
        
        lista_receta[elem.idReceta].append(elem)

    agrupado_dict = dict(lista_receta)
    list_productos = []

    for id_receta, items in lista_receta.items():

        receta = {
            "nombre_receta": None,
            "capacidad_receta": 0, 
            "cantidad_ciclos": 0, 
        }
        db_receta = (
            db.query(Receta)
            .filter(Receta.id == id_receta)
            .first())
        for item in items:
            receta["nombre_receta"] = db_receta.nombre
            receta["capacidad_receta"] += int(item.peso)
            receta["cantidad_ciclos"] += 1

        list_productos.append(receta)

    respuesta["ciclos_realizados"] = cantidad_ciclos_incorectos + cantidad_ciclos_correctos
    respuesta["produccion_total"] = round(total_peso / 1000, 2)
    respuesta["ciclos_correctos"] = cantidad_ciclos_correctos
    respuesta["ciclos_incorrectos"] = cantidad_ciclos_incorectos
    respuesta["productos_realizados"] = list_productos
 
    return respuesta

def obtener_valor_sensores(db, id_ciclo:int, id_sensor:int, tipo):
    if tipo == "MAX":
        resultado = (
            db.query(SensoresAA)
            .filter(SensoresAA.idCiclo == id_ciclo, SensoresAA.idSensor == id_sensor)
            .order_by(SensoresAA.valor.desc())
            .limit(1)
            .first()
        )
        return resultado.valor
    if tipo == "MIN":
        resultado = (
            db.query(SensoresAA)
            .filter(SensoresAA.idCiclo == id_ciclo, SensoresAA.idSensor == id_sensor)
            .order_by(SensoresAA.valor.asc())
            .limit(1)
            .first()
        )
        return resultado.valor
    return 0

def obtener_datos_graficos(db, id_ciclo:int):
    lista_sensores_data = {}

    ciclo: list[Ciclo] = []
    if id_ciclo != 0: 
        ciclo = (db.query(Ciclo)
                .filter(Ciclo.id == id_ciclo)
                .first()
                )
    if id_ciclo == 0:
        ciclo = db.query(Ciclo).order_by(Ciclo.id.desc()).first()  

    if not ciclo:
        logger.info("No se encontró el ciclo en la BDD")
        raise ValueError("Ciclo no encontrado")


    lista_sensores = (
        db.query(Sensores)
    ).all()

    estado_ciclo = (
        db.query(EstadoCiclo)
        .filter(EstadoCiclo.idCiclo == ciclo.id)
    )
    receta = (
        db.query(Receta)
        .filter(ciclo.idReceta == Receta.id)
        .first()
    )

    general = {}
    general["id_ciclo"] = ciclo.id
    general["ciclo_lote"] = ciclo.lote
    general["tiempo_transcurrido"] = ciclo.tiempoTranscurrido
    general["fecha_inicio"] = ciclo.fecha_inicio
    general["fecha_fin"] = ciclo.fecha_fin
    general["receta"] = receta.nombre

    for sensor in lista_sensores:
        if sensor.nombre == "Temperatura agua":
            temp_agua = (
                db.query(SensoresAA)
                .filter(SensoresAA.idSensor == sensor.id)
                .filter(SensoresAA.idCiclo == ciclo.id)
                .all()
            )
            print(f"Cantidad Filas de registros en BDD: {len(temp_agua)}")
            general["temp_agua_max"] = obtener_valor_sensores(db,id_ciclo, sensor.id, "MAX")
            general["temp_agua_min"] = obtener_valor_sensores(db,id_ciclo, sensor.id, "MIN")
            lista_sensores_data[sensor.nombre] = temp_agua
        
        if sensor.nombre == "Temperatura producto":
            temp_producto = (
                db.query(SensoresAA)
                .filter(SensoresAA.idSensor == sensor.id)
                .filter(SensoresAA.idCiclo == ciclo.id)
                .all()
            )
            print(f"Cantidad Filas de registros en BDD: {len(temp_producto)}")
            general["temp_producto_max"] = obtener_valor_sensores(db,id_ciclo, sensor.id, "MAX")
            general["temp_producto_min"] = obtener_valor_sensores(db,id_ciclo, sensor.id, "MIN")
            lista_sensores_data[sensor.nombre] = temp_producto
        if sensor.nombre == "Nivel agua":
            nivel_agua = (
                db.query(SensoresAA)
                .filter(SensoresAA.idSensor == sensor.id)
                .filter(SensoresAA.idCiclo == ciclo.id)
                .all()
            )
            print(f"Cantidad Filas de registros en BDD: {len(nivel_agua)}")
            general["nivel_agua_max"] = obtener_valor_sensores(db,id_ciclo, sensor.id, "MAX")
            general["nivel_agua_min"] = obtener_valor_sensores(db,id_ciclo, sensor.id, "MIN")

            lista_sensores_data[sensor.nombre] = nivel_agua

    lista_sensores_data["general"] = general

    return lista_sensores_data

def generar_informe_ciclo(db, id_ciclo:int, equipo:str):
    ciclo: list[Ciclo] = []
    if id_ciclo != 0: 
        ciclo = (db.query(Ciclo)
                .filter(Ciclo.id == id_ciclo)
                .first()
                )
    if id_ciclo == 0:
        ciclo = db.query(Ciclo).order_by(Ciclo.id.desc()).first()  

    if not ciclo:
        logger.info("No se encontró el ciclo en la BDD")
        raise ValueError("Ciclo no encontrado")


    lista_sensores = (
        db.query(Sensores)
    ).all()

    estado_ciclo = (
        db.query(EstadoCiclo)
        .filter(EstadoCiclo.idCiclo == ciclo.id)
    )

    receta = (
        db.query(Receta)
        .filter(ciclo.idReceta == Receta.id)
        .first()
    )

    maquina = (
        db.query(Equipo)
        .filter(ciclo.idEquipo == Equipo.id)
        .first()
    )

    cantidad_filas = 0

    temp_agua: list[SensoresAA] = []
    temp_producto: list[SensoresAA] = []
    nivel_agua: list[SensoresAA] = []

    for sensor in lista_sensores:
        if sensor.nombre == "Temperatura agua":
            temp_agua = (
                db.query(SensoresAA)
                .filter(SensoresAA.idSensor == sensor.id)
                .filter(SensoresAA.idCiclo == ciclo.id)
                .all()
            )
            print(f"Cantidad Filas de registros en BDD: {len(temp_agua)}")
            if len(temp_agua)>= cantidad_filas:
                cantidad_filas = len(temp_agua)
        
        if sensor.nombre == "Temperatura producto":
            temp_producto = (
                db.query(SensoresAA)
                .filter(SensoresAA.idSensor == sensor.id)
                .filter(SensoresAA.idCiclo == ciclo.id)
                .all()
            )
            print(f"Cantidad Filas de registros en BDD: {len(temp_producto)}")
            if len(temp_producto)>= cantidad_filas:
                cantidad_filas = len(temp_producto)
        if sensor.nombre == "Nivel agua":
            nivel_agua = (
                db.query(SensoresAA)
                .filter(SensoresAA.idSensor == sensor.id)
                .filter(SensoresAA.idCiclo == ciclo.id)
                .all()
            )
            print(f"Cantidad Filas de registros en BDD: {len(nivel_agua)}")
            if len(nivel_agua)>= cantidad_filas:
                cantidad_filas = len(nivel_agua)


    print(f"VALOR MAS ALTO ELEMENTOS DE SENSORES: {cantidad_filas}")

    #CONSTRUCCION DEL ARCHIVO XLMS
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Informe de Ciclo de Coccion"
    logoPath = "cremona.png"
    img = Image(logoPath)
    img.width = 280
    img.height = 70
    sheet.add_image(img, 'D1')

    sheet.append([f"Receta: {receta.nombre}"])
    producto_cell = sheet.cell(row=sheet.max_row, column=1)
    producto_cell.font = Font(bold= True, size=20)

    sheet.append([f"Peso del producto: {ciclo.peso}kg"])
    producto_cell = sheet.cell(row=sheet.max_row, column=3)
    producto_cell.font = Font(bold= True, size=16)

    sheet.append([f"N° de lote: {ciclo.lote}"])
    producto_cell = sheet.cell(row=sheet.max_row, column=4)
    producto_cell.font = Font(bold= True, size=16)

    sheet.append([f"Tiempo Transcurrido: {ciclo.tiempoTranscurrido}hs:mm"])
    producto_cell = sheet.cell(row=sheet.max_row, column=5)
    producto_cell.font = Font(bold= True, size=16)

    headers = [
        "ID_CICLO", "EQUIPO", "LOTE", "ESTADO CICLO", "NOMBRE RECETA",
        "TEMPERARURA AGUA  [°C]", "TEMPERATURA PRODUCTO [°C]", "NIVEL AGUA [mm]", "FECHA REGISTRO"
    ]
    sheet.append(headers)
    header_fill = PatternFill(start_color="145f82", end_color="145f82", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)  

    for col in range(1, len(headers) + 1):
        cell = sheet.cell(row=sheet.max_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    start_row = sheet.max_row + 1

    resultado_fila = []

    for i, elem in enumerate(temp_producto):
        resultado_fila.append([
            ciclo.id,
            maquina.nombre, 
            ciclo.lote, 
            "Finalizado INC",
            receta.nombre,
            temp_agua[i].valor, 
            elem.valor, 
            nivel_agua[i].valor, 
            elem.fechaRegistro
        ])

    for item in resultado_fila:
        sheet.append(item)
    
    end_row = sheet.max_row
    start_col = 1
    end_col = len(headers)
    table_range = f"{sheet.cell(row=start_row -1, column=start_col).coordinate}:{sheet.cell(row=end_row, column=end_col).coordinate}"
    table_nombre = "GraficosHistorico"
    tabla = Table(displayName=table_nombre, ref=table_range)
    style = TableStyleInfo(showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tabla.tableStyleInfo = style

    # Agregar la tabla a la hoja
    sheet.add_table(tabla)
    sheet.append([])

    for col in sheet.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        sheet.column_dimensions[column_letter].width = max_length + 2

    excel_stream = BytesIO()
    workbook.save(excel_stream)
    workbook.close() 
    excel_stream.seek(0)  

    return excel_stream